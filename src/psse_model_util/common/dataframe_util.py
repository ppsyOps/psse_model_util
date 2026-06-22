import re
from datetime import datetime as dtdt
from pathlib import Path
from typing import Any, Callable, List, Tuple, Union

import arrow
import numpy as np
import openpyxl
import pandas as pd
import pytz
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from psse_model_util.dataformat.classes import ModelDF


def convert_df_column_dtypes(
        df_in: pd.DataFrame | ModelDF,
        new_dtypes: dict = None,
        convert_all_columns: bool = True,
        inplace: bool = False,
        default_types: Tuple[Union[Callable, type]] = (pd.to_datetime, int, float),
) -> pd.DataFrame | ModelDF:
    """
    Converts the column types of a dataframe from str to datetime, float, or int if possible.

    :param df_in: The dataframe to convert.
    :param new_dtypes: A dictionary of {column_name: new_dtype},
                       where column_name is the name of the column and
                       new_dtype is an ordered list of datatypes to try to convert
                       the column to. If None is provided, defaults are
                       [pd.to_datetime, float, int].
    :param convert_all_columns: If True, any columns in df but not in
                                new_dtypes will be converted using default types.
                                If False, only columns specified in new_dtypes
                                will be converted.
    :param inplace: If True, modify the original dataframe. If False,
                    return a new dataframe.
    :param default_types: If new_dtypes not specified at or or not for a
                          specific column, then default_types will be used for
                          the conversion.
    :return: The dataframe with converted column types.

    :raises AssertionError: If df is not a pandas DataFrame.
    """
    assert isinstance(df_in, (pd.DataFrame, ModelDF)), \
        (f"Expected df_in to be a pandas DataFrame; got {type(df_in)}.")

    # default_types: Tuple[Union[Callable, type]] = pd.to_datetime, int, float
    new_dtypes = new_dtypes or {}
    metadata = df_in._metadata
    if isinstance(df_in, ModelDF):
        meta = df_in.meta

    df_out = df_in if inplace else df_in.copy()

    def convert_column(column: pd.Series,
                       try_dtypes: List[Union[Callable[[Any], Any], type]] = default_types,
                       ) -> pd.Series:
        """Try to convert a column to specified types, otherwise keep as original."""
        # print(f"convert_column(column = {column.name}")
        # print(f"                        {column}")
        # print(f"               try_dtypes = {try_dtypes})")
        if isinstance(try_dtypes, type):
            try_dtypes = [try_dtypes]
        for dtype in try_dtypes:
            # The approach below allows for a wide range of successful
            # conversions, as it combines the efficiency of astype() with the
            # flexibility of apply(). It can handle cases where some elements
            # in a column might require special treatment that astype() can't
            # provide.
            try:
                if dtype == pd.to_datetime:
                    # For datetime, it uses pd.to_datetime(col), which is the
                    # most appropriate method for datetime conversion.
                    return pd.to_datetime(column)
                elif dtype is int or (isinstance(dtype, type) and issubclass(dtype, int)):
                    # Try to convert to int, but only if it doesn't result in data loss
                    float_series = column.astype(float)
                    if (float_series == float_series.round()).all():
                        return float_series.astype(int)
                    else:
                        continue
                else:
                    # Attempt col.astype(dtype), which is generally faster and
                    # works for many straightforward conversions.
                    return column.astype(dtype)
            except (ValueError, TypeError):
                try:
                    # Fallback conversion attempt. If the primary conversion
                    # fails, it then tries col.apply(dtype). This is
                    # particularly useful for custom conversion functions
                    # or when individual element conversion is necessary.
                    return column.apply(dtype)
                except (ValueError, TypeError):
                    # Both conversion methods failed. Continue to the next
                    # dtype in the list.
                    continue
        return column

    # If convert_all_columns = True, then ensure comprehensive coverage of all
    # columns when requested, even if they weren't explicitly mentioned in new_dtypes.
    if convert_all_columns:
        for column_name in df_out.columns:
            if column_name not in new_dtypes:
                new_dtypes[column_name] = default_types

    # Convert columns in new_dtypes
    for column_name, types in new_dtypes.items():
        if column_name in df_out.columns:  # Check if the column exists in the dataframe
            # The use of types = types or default_types allows for flexible
            # type specification. Users can provide specific types for some
            #  columns and rely on defaults for others.
            types = types or default_types
            if isinstance(types, type):
                types = [types]
            df_out[column_name] = convert_column(df_out[column_name], try_dtypes=types)

    if hasattr(df_in, 'meta'):
        df_out.meta = meta
    df_out._metadata = metadata

    return df_out


# If df[col1] is null get value from df[col1]
def coalesce(df: pd.DataFrame, new_col: str, col1: str, col2: str,
             drop=True) -> pd.DataFrame:
    """
    caolesce says: if data in one column (col1) of a dataframe is missing, then fill
    it with data from another column (col2) and save it to new_col.
    :param df: pd.DataFrame to process
    :param new_col: name of column in df in which to place the result.  Note
                    that you could set new_col to an existing column name, such
                    as col1, which would overwrite that column.
    :param col1: Use data from this column if not None/NaN, else use col2
    :param col2: Use data from this column if col1 contains None/NaN
    :param drop: bool: if True, drop col1 and col2.
                       if True, do not drop any columns.
    :return: pd.DataFrame
    Make column c which takes the value in c1, unless it is null,
    in which case it uses the value in c2.  If `drop`, columns c1 and
    c2 and removed"""
    df[new_col] = df[col1].fillna(df[col2])
    if drop:
        df.drop([col1, col2], axis=1, inplace=True)
    return df


def create_empty_DataFrame(columns, index_col=None):
    """
    Create an empty dataframe and specify the datatype of each column.
    Ex: df = create_empty_DataFrame([('my_date', 'datetime64[ns]'),
                                     ('my_num', int),
                                     ('id', str),
                                     ('primary', bool),
                                     ('side', str),
                                     ('quantity', int),
                                     ('price', float)
                                     ]
                                    )
    :param columns: list of 2-tuples, where each tuple is a pair of column name and data type.
    :return: an empty pandas DataFrame with the specified column names and data types.
    """
    if index_col:
        index_type = next((t for name, t in columns if name == index_col))
        df = pd.DataFrame({name: pd.Series(dtype=t)
                           for name, t in columns if name != index_col},
                          index=pd.Index([], dtype=index_type))
    else:
        df = pd.DataFrame({name: pd.Series(dtype=t)
                           for name, t in columns if name != index_col})
    cols = [name for name, _ in columns]
    if index_col:
        cols.remove(index_col)
    return df[cols]


def make_naive(datetime_value: arrow.Arrow | dtdt,
               as_tz=None):
    """
    Converts an arrow.Arrow or datetime with timezone to a datetime
    that is time zone naive.  If a timezone is part of datetime_value,
    the datetime is converted to as_tz before stripping time zone info.
    :param datetime_value:
    :return: naive datetime.datetime value, i.e., datetime_value stripped of
             tz_info.
    """
    if datetime_value.tzinfo is None:
        return datetime_value
    elif isinstance(datetime_value, arrow.Arrow):
        datetime_value = datetime_value.datetime
    if as_tz:
        return datetime_value.astimezone(as_tz).replace(tzinfo=None)
    else:
        return datetime_value.replace(tzinfo=None)


def make_df_naive(df: pd.DataFrame, as_tz=pytz.UTC):
    """
    Convert all datetime columns in a dataframe from aware to naive.
    :param df: pd.DataFrame to process.
    :param as_tz: a valid datetime.tzinfo object, like pytz.UTC
    :return: DataFrame with naive datetime columns
    """
    df = df.copy()  # Create a copy to avoid modifying the original DataFrame
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            if df[col].dt.tz is not None:
                df[col] = df[col].dt.tz_convert(as_tz).dt.tz_localize(None)
            else:
                df[col] = df[col].dt.tz_localize(as_tz).dt.tz_localize(None)
    return df


def df_column_validator(df_column: pd.Series,
                        nullable: bool = None,
                        col_dtype=None,
                        col_min=None,
                        col_max=None,
                        isin: list = [],
                        pattern: str = '',
                        pattern_criteria: str = 'all') -> str:
    pattern_criteria: str = pattern_criteria.lower()
    if not isinstance(df_column, pd.Series) \
            and not isinstance(df_column, pd.DataFrame):
        df_column = pd.Series(df_column)
    result = ''
    if nullable is False and df_column.isnull().values.any():
        result += 'Cannot contain null values.  '
    if col_dtype:
        if isinstance(col_dtype, (list, tuple)):
            if df_column.dtype not in [np.dtype(dt) for dt in col_dtype]:
                result += f'Expected dtype to be in {col_dtype}; got {df_column.dtype}.  '
        elif np.dtype(col_dtype) != df_column.dtype:
            if 'datetime' in str(col_dtype) and 'datetime' in str(df_column.dtype):
                # df_column is a datetime as expected.
                pass
            elif (col_dtype is int or (isinstance(col_dtype, type) and issubclass(col_dtype, int))) and df_column.dtype == np.int64:
                # Special case for int vs int64
                pass
            else:
                result += f'Expected dtype {col_dtype}; got {df_column.dtype}.  '
    if col_min is not None and df_column.min() < col_min:
        result += f'{df_column.min()} < allowable min, {col_min}.  '
    if col_max is not None and df_column.max() > col_max:
        result += f'{df_column.max()} > allowable max, {col_max}.  '
    if isin and not df_column.isin(isin).all():
        result += f'Contains a value not in: {isin}.  '
    if pattern:
        series = df_column.astype(str, copy=True)
        if isinstance(pattern, str):
            compiled = re.compile(pattern)
        elif isinstance(pattern, re.Pattern):
            compiled = pattern
        else:
            raise TypeError('pattern must be str or re.Pattern (i.e., a '
                            'compiled regex pattern).')

        if pattern_criteria == 'all':
            b = series.apply(lambda x: bool(re.search(pattern=compiled,
                                                      string=x))).all()
        elif pattern_criteria == 'any':
            b = series.apply(lambda x: bool(re.search(pattern=compiled,
                                                      string=x))).any()
        else:
            raise ValueError('pattern_criteria must be in ["all", "any"]')
        if not b:
            result += f're.search({pattern}, ["{df_column.name}"]).{pattern_criteria}() not satisfied.  '

    return result.strip()


def df_to_excel_worksheet(
    dataframe: pd.DataFrame,
    sheet_name: str,
    filepath: Union[str, Path] = Path(),
    if_sheet_exists: str = 'replace',
    chunk_size: int = 100000,
    make_naive: bool = False,
    **kwargs
) -> Path:
    """
    Write a dataframe to a specific Excel worksheet without overwriting other
    worksheets, optimized for speed.

    Args:
        dataframe (pd.DataFrame): The dataframe to export to Excel.
        sheet_name (str): The name of the Excel worksheet to write to.
        filepath (Union[str, Path]): The path to the Excel workbook.
        if_sheet_exists (str): Action to take if the sheet already exists.
            Options are 'replace', 'overlay', or 'new'.
        chunk_size (int): Number of rows to write at a time for large datasets.
        make_naive (bool): Whether to convert datetime columns to naive datetimes.
        **kwargs: Additional arguments to pass to openpyxl's worksheet.append() method.

    Returns:
        Path: The path to the created or modified Excel file.

    Raises:
        ValueError: If an invalid option is provided for if_sheet_exists.
    """
    filepath = Path(filepath)
    if not filepath.parent.exists():
        filepath.parent.mkdir(parents=True, exist_ok=True)

    # Convert datetime columns to naive if requested
    if make_naive:
        dataframe = _make_df_naive(dataframe)

    # Use openpyxl directly for faster writing
    if filepath.exists():
        workbook = openpyxl.load_workbook(filepath)
        if if_sheet_exists == 'replace' and sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
            worksheet = workbook.create_sheet(sheet_name)
        elif if_sheet_exists == 'overlay' and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        elif if_sheet_exists == 'new' or sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
        else:
            raise ValueError("Invalid option for if_sheet_exists. Choose 'replace', 'overlay', or 'new'.")
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    # Write data in chunks
    _write_dataframe_to_worksheet(dataframe, worksheet, chunk_size, **kwargs)

    # Save the workbook
    workbook.save(filepath)

    return filepath

def _write_dataframe_to_worksheet(
    df: pd.DataFrame,
    worksheet: Worksheet,
    chunk_size: int = 100000,
    **kwargs
) -> None:
    """
    Write a DataFrame to an openpyxl worksheet in chunks.

    Args:
        df (pd.DataFrame): The DataFrame to write.
        worksheet (Worksheet): The openpyxl worksheet to write to.
        chunk_size (int): Number of rows to write at a time.
        **kwargs: Additional arguments to pass to worksheet.append().
    """
    # Write headers
    headers = list(df.columns)
    worksheet.append(headers)
    kwargs.setdefault('index', False)
    kwargs.setdefault('header', False)
    index = kwargs.pop('index')
    header = kwargs.pop('header')
    # Write data in chunks
    for i in range(0, len(df), chunk_size):
        chunk = df.iloc[i:i+chunk_size]
        for row in dataframe_to_rows(chunk, index=index, header=header):
            worksheet.append(row, **kwargs)

def _make_df_naive(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert all datetime columns in a dataframe from aware to naive.

    Args:
        df (pd.DataFrame): The input DataFrame.

    Returns:
        pd.DataFrame: DataFrame with naive datetime columns.
    """
    for col in df.columns:
        s = df[col]
        # select_dtypes(include=['datetime64']) matches only tz-NAIVE columns,
        # so the old version skipped the tz-aware columns it was meant to fix
        # (and would have raised on the naive ones). Strip tz from aware columns
        # only; leave naive datetimes and non-datetimes untouched.
        if pd.api.types.is_datetime64_any_dtype(s) and s.dt.tz is not None:
            df[col] = s.dt.tz_localize(None)
    return df


def move_columns_far_left(df: pd.DataFrame, columns_to_move: list | tuple):
    """
    Move a specified subset of columns (columns_to_move) to the far left (the beginning) of the dataframe.

    Verbose version of code:
        # Filter out cols_to_move from the original DataFrame's columns
        remaining_cols = [col for col in df.columns if col not in columns_to_move]
        # Concatenate the moved columns with the remaining columns
        new_order = columns_to_move + remaining_cols
        #  Reindex the DataFrame using this new column order
        return df[new_order]

    :param df: pd.DataFrame for which to modify the column order
    :param columns_to_move: Columns to move to the far left (the beginning) of the dataframe.
    :return: The DataFrame with columns re-ordered.
    """
    if df.empty and len(df.columns) == 0:
        return df

    # Filter out non-existent columns
    existing_columns_to_move = [col for col in columns_to_move if col in df.columns]

    new_order = existing_columns_to_move + [col for col in df.columns if col not in existing_columns_to_move]
    return df.reindex(columns=new_order)


def move_columns_far_right(df: pd.DataFrame, columns_to_move: list | tuple):
    """
    Move a specified subset of columns (columns_to_move) to the far right (the end) of the dataframe.
    :param df: pd.DataFrame for which to modify the column order
    :param columns_to_move: Columns to move to the far right (the end) of the dataframe.
    :return: The DataFrame with columns re-ordered.
    """
    if df.empty and len(df.columns) == 0:
        return df

    # Filter out non-existent columns
    existing_columns_to_move = [col for col in columns_to_move if col in df.columns]

    new_order = [col for col in df.columns if col not in existing_columns_to_move] + existing_columns_to_move
    return df.reindex(columns=new_order)


if __name__ == '__main__':
    # --------------  Export Large Dataframe to Excel--------------------------

    from pathlib import Path

    import pandas as pd

    # Create a sample DataFrame
    df = pd.DataFrame({'A': range(1000000), 'B': range(1000000, 2000000)})

    # Export to Excel
    filepath = Path('large_file.xlsx')
    df_to_excel_worksheet(df, 'Sheet1', filepath, chunk_size=100000)

    # # --------------------  Multiple Dataframes -------------------------------
    #
    # # SQL-like join two dataframes (inner, outer, left, right, cross).
    # # Note: use now="cross" for a cartesian join.
    # df1 = pd.DataFrame({'A': list(range(10)), 'Ax2': list(range(0, 20, 2))})
    # df2 = pd.DataFrame({'A': list(range(10)), 'Ax3': list(range(0, 30, 3))})
    # df3 = pd.merge(df1, df2, how='left', left_on=['A'], right_on=['A'])
    #
    # # --------------------------  Columns  ------------------------------------
    #
    # # Reorder columns.
    # df = pd.DataFrame(np.random.rand(10, 4), columns=['A', 'C', 'B', 'D'])
    # df = df.reindex(columns=['A', 'B', 'C', 'D'])
    #
    # # Drop columns.
    # df = pd.DataFrame(np.random.rand(10, 4), columns=['A', 'B', 'C', 'D'])
    # df.drop(['C', 'D'], axis=1)
