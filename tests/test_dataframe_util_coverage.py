"""
test_dataframe_util_coverage.py — characterization tests to raise coverage of
``psse_model_util.common.dataframe_util``.

These tests exercise the pure validation / IO helpers only. They deliberately
avoid loading ``Model`` objects or touching the pickle cache.
"""
from __future__ import annotations

import re

import numpy as np
import openpyxl
import pandas as pd
import pytest

from psse_model_util.common import dataframe_util as dfu

# ---------------------------------------------------------------------------
# convert_df_column_dtypes
# ---------------------------------------------------------------------------

def test_convert_df_column_dtypes_int_float_datetime():
    df = pd.DataFrame(
        {
            "i": ["1", "2", "3"],
            "f": ["1.5", "2.5", "3.5"],
            "d": ["2020-01-01", "2020-01-02", "2020-01-03"],
            "s": ["a", "b", "c"],
        }
    )
    out = dfu.convert_df_column_dtypes(df)
    assert pd.api.types.is_integer_dtype(out["i"])
    assert pd.api.types.is_float_dtype(out["f"])
    assert pd.api.types.is_datetime64_any_dtype(out["d"])
    # str column is not numeric/datetime; it remains a (string-like) object column
    assert not pd.api.types.is_numeric_dtype(out["s"])
    assert not pd.api.types.is_datetime64_any_dtype(out["s"])
    # original untouched (inplace=False): "i" stays non-integer in the source df
    assert not pd.api.types.is_integer_dtype(df["i"])
    assert out is not df


def test_convert_df_column_dtypes_inplace_returns_same_object():
    df = pd.DataFrame({"i": ["1", "2"]})
    out = dfu.convert_df_column_dtypes(df, inplace=True)
    assert out is df


def test_convert_df_column_dtypes_new_dtypes_scalar_type():
    # Passing a bare type (not a list) for a column exercises the
    # isinstance(types, type) -> [types] normalization branch.
    df = pd.DataFrame({"x": ["1", "2", "3"]})
    out = dfu.convert_df_column_dtypes(
        df, new_dtypes={"x": float}, convert_all_columns=False
    )
    assert pd.api.types.is_float_dtype(out["x"])


def test_convert_df_column_dtypes_non_round_floats_stay_float():
    # int conversion is skipped when rounding would lose data.
    df = pd.DataFrame({"x": ["1.5", "2.5"]})
    out = dfu.convert_df_column_dtypes(df)
    assert pd.api.types.is_float_dtype(out["x"])


def test_convert_df_column_dtypes_assert_non_dataframe():
    with pytest.raises(AssertionError):
        dfu.convert_df_column_dtypes([1, 2, 3])


# ---------------------------------------------------------------------------
# coalesce
# ---------------------------------------------------------------------------

def test_coalesce_drop_true():
    df = pd.DataFrame({"a": [1.0, np.nan, 3.0], "b": [10.0, 20.0, 30.0]})
    out = dfu.coalesce(df, "c", "a", "b", drop=True)
    assert list(out["c"]) == [1.0, 20.0, 3.0]
    assert "a" not in out.columns
    assert "b" not in out.columns


def test_coalesce_drop_false():
    df = pd.DataFrame({"a": [1.0, np.nan], "b": [10.0, 20.0]})
    out = dfu.coalesce(df, "c", "a", "b", drop=False)
    assert "a" in out.columns and "b" in out.columns


# ---------------------------------------------------------------------------
# create_empty_DataFrame
# ---------------------------------------------------------------------------

def test_create_empty_dataframe_no_index():
    df = dfu.create_empty_DataFrame(
        [("name", str), ("num", int), ("price", float)]
    )
    assert list(df.columns) == ["name", "num", "price"]
    assert len(df) == 0
    assert pd.api.types.is_integer_dtype(df["num"])
    assert pd.api.types.is_float_dtype(df["price"])


def test_create_empty_dataframe_with_index_col():
    # Exercises lines 166-167 (index branch) and 175 (cols.remove).
    df = dfu.create_empty_DataFrame(
        [("id", "int64"), ("val", float)], index_col="id"
    )
    assert df.index.name is None  # index column dropped from columns
    assert list(df.columns) == ["val"]
    assert "id" not in df.columns
    assert len(df) == 0


# ---------------------------------------------------------------------------
# make_naive / make_df_naive
# ---------------------------------------------------------------------------

def test_make_naive_already_naive():
    import datetime as _dt

    naive = _dt.datetime(2020, 1, 1, 12, 0, 0)
    assert dfu.make_naive(naive) is naive


def test_make_naive_aware_datetime():
    import datetime as _dt

    aware = _dt.datetime(2020, 1, 1, 12, 0, 0, tzinfo=_dt.timezone.utc)
    out = dfu.make_naive(aware)
    assert out.tzinfo is None


def test_make_naive_arrow_with_as_tz():
    import datetime as _dt

    import arrow

    a = arrow.Arrow(2020, 1, 1, 12, 0, 0, tzinfo="UTC")
    out = dfu.make_naive(a, as_tz=_dt.timezone.utc)
    assert out.tzinfo is None


def test_make_df_naive_aware_and_naive_columns():
    aware = pd.to_datetime(
        ["2020-01-01 00:00", "2020-01-02 00:00"]
    ).tz_localize("US/Eastern")
    naive = pd.to_datetime(["2021-01-01", "2021-01-02"])
    df = pd.DataFrame({"aware": aware, "naive": naive, "other": [1, 2]})
    out = dfu.make_df_naive(df)
    assert out["aware"].dt.tz is None
    assert out["naive"].dt.tz is None
    # original not mutated
    assert df["aware"].dt.tz is not None


# ---------------------------------------------------------------------------
# df_column_validator
# ---------------------------------------------------------------------------

def test_validator_passes_clean():
    s = pd.Series([1, 2, 3])
    assert dfu.df_column_validator(s) == ""


def test_validator_coerces_non_series_input():
    # Line 227: a plain list gets wrapped in pd.Series.
    msg = dfu.df_column_validator([1, 2, 3], nullable=False)
    assert msg == ""


def test_validator_nullable_false_with_nulls():
    s = pd.Series([1.0, np.nan, 3.0])
    msg = dfu.df_column_validator(s, nullable=False)
    assert "Cannot contain null values" in msg


def test_validator_col_dtype_scalar_mismatch():
    s = pd.Series(["a", "b"])
    msg = dfu.df_column_validator(s, col_dtype=float)
    assert "Expected dtype" in msg


def test_validator_col_dtype_list_mismatch():
    s = pd.Series(["a", "b"])  # object dtype
    msg = dfu.df_column_validator(s, col_dtype=[np.int64, np.float64])
    assert "Expected dtype to be in" in msg


def test_validator_col_dtype_list_match():
    s = pd.Series([1, 2, 3], dtype="int64")
    msg = dfu.df_column_validator(s, col_dtype=[np.int64, np.float64])
    assert msg == ""


def test_validator_col_dtype_datetime_special_case():
    s = pd.to_datetime(pd.Series(["2020-01-01", "2020-01-02"]))
    # Expecting datetime when column is datetime: the special-case passes.
    msg = dfu.df_column_validator(s, col_dtype="datetime64[ns]")
    assert msg == ""


def test_validator_col_dtype_int_vs_int64_special_case():
    s = pd.Series([1, 2, 3], dtype="int64")
    # col_dtype int vs int64 column: special case passes.
    msg = dfu.df_column_validator(s, col_dtype=int)
    assert msg == ""


def test_validator_col_min():
    s = pd.Series([1, 2, 3])
    msg = dfu.df_column_validator(s, col_min=2)
    assert "allowable min" in msg


def test_validator_col_max():
    s = pd.Series([1, 2, 3])
    msg = dfu.df_column_validator(s, col_max=2)
    assert "allowable max" in msg


def test_validator_isin_violation():
    s = pd.Series(["x", "y", "z"])
    msg = dfu.df_column_validator(s, isin=["x", "y"])
    assert "Contains a value not in" in msg


def test_validator_isin_ok():
    s = pd.Series(["x", "y"])
    msg = dfu.df_column_validator(s, isin=["x", "y", "z"])
    assert msg == ""


def test_validator_pattern_str_all_pass():
    s = pd.Series(["a1", "a2", "a3"], name="col")
    msg = dfu.df_column_validator(s, pattern=r"a\d", pattern_criteria="all")
    assert msg == ""


def test_validator_pattern_str_all_fail():
    s = pd.Series(["a1", "bb", "a3"], name="col")
    msg = dfu.df_column_validator(s, pattern=r"a\d", pattern_criteria="all")
    assert "not satisfied" in msg


def test_validator_pattern_any_pass():
    s = pd.Series(["xx", "a1", "yy"], name="col")
    msg = dfu.df_column_validator(s, pattern=r"a\d", pattern_criteria="any")
    assert msg == ""


def test_validator_pattern_any_fail():
    s = pd.Series(["xx", "yy"], name="col")
    msg = dfu.df_column_validator(s, pattern=r"a\d", pattern_criteria="any")
    assert "not satisfied" in msg


def test_validator_pattern_compiled():
    s = pd.Series(["a1", "a2"], name="col")
    compiled = re.compile(r"a\d")
    msg = dfu.df_column_validator(s, pattern=compiled, pattern_criteria="all")
    assert msg == ""


def test_validator_pattern_invalid_criteria_raises():
    s = pd.Series(["a1", "a2"], name="col")
    with pytest.raises(ValueError):
        dfu.df_column_validator(s, pattern=r"a\d", pattern_criteria="bogus")


def test_validator_pattern_wrong_type_raises():
    s = pd.Series(["a1", "a2"], name="col")
    with pytest.raises(TypeError):
        dfu.df_column_validator(s, pattern=123)


def test_validator_combined_messages():
    s = pd.Series([1.0, np.nan, 100.0])
    msg = dfu.df_column_validator(s, nullable=False, col_max=50)
    assert "Cannot contain null values" in msg
    assert "allowable max" in msg


# ---------------------------------------------------------------------------
# df_to_excel_worksheet / _write_dataframe_to_worksheet / _make_df_naive
# ---------------------------------------------------------------------------

def _read_sheet(path, sheet_name):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def test_excel_create_new_file(tmp_path):
    df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
    fp = tmp_path / "out.xlsx"
    result = dfu.df_to_excel_worksheet(df, "Data", fp)
    assert result == fp
    assert fp.exists()
    rows = _read_sheet(fp, "Data")
    assert rows[0] == ["a", "b"]
    assert rows[1] == [1, 3]
    assert rows[2] == [2, 4]


def test_excel_create_parent_dir(tmp_path):
    df = pd.DataFrame({"a": [1]})
    fp = tmp_path / "nested" / "sub" / "out.xlsx"
    dfu.df_to_excel_worksheet(df, "S", fp)
    assert fp.exists()


def test_excel_replace_existing_sheet(tmp_path):
    fp = tmp_path / "out.xlsx"
    dfu.df_to_excel_worksheet(pd.DataFrame({"a": [1, 2]}), "Sheet1", fp)
    # Replace it with different data.
    dfu.df_to_excel_worksheet(
        pd.DataFrame({"z": [9]}), "Sheet1", fp, if_sheet_exists="replace"
    )
    rows = _read_sheet(fp, "Sheet1")
    assert rows[0] == ["z"]
    assert rows[1] == [9]


def test_excel_overlay_existing_sheet(tmp_path):
    fp = tmp_path / "out.xlsx"
    dfu.df_to_excel_worksheet(pd.DataFrame({"a": [1]}), "Sheet1", fp)
    # Overlay appends onto the existing sheet (does not delete it).
    dfu.df_to_excel_worksheet(
        pd.DataFrame({"a": [2]}), "Sheet1", fp, if_sheet_exists="overlay"
    )
    rows = _read_sheet(fp, "Sheet1")
    # original header + row, then overlaid header + row appended
    assert len(rows) == 4


def test_excel_new_sheet_branch(tmp_path):
    fp = tmp_path / "out.xlsx"
    dfu.df_to_excel_worksheet(pd.DataFrame({"a": [1]}), "Sheet1", fp)
    dfu.df_to_excel_worksheet(
        pd.DataFrame({"b": [2]}), "Sheet2", fp, if_sheet_exists="new"
    )
    wb = openpyxl.load_workbook(fp)
    assert "Sheet1" in wb.sheetnames
    assert "Sheet2" in wb.sheetnames


def test_excel_invalid_if_sheet_exists_raises(tmp_path):
    fp = tmp_path / "out.xlsx"
    dfu.df_to_excel_worksheet(pd.DataFrame({"a": [1]}), "Sheet1", fp)
    with pytest.raises(ValueError):
        dfu.df_to_excel_worksheet(
            pd.DataFrame({"a": [2]}), "Sheet1", fp, if_sheet_exists="bogus"
        )


def test_excel_make_naive_branch_runs(tmp_path):
    # Exercises the `if make_naive: dataframe = _make_df_naive(...)` branch with
    # a naive datetime column (which openpyxl can write).
    naive = pd.to_datetime(["2020-01-01", "2020-01-02"])
    df = pd.DataFrame({"ts": naive, "v": [1, 2]})
    fp = tmp_path / "out.xlsx"
    dfu.df_to_excel_worksheet(df, "Data", fp, make_naive=True)
    assert fp.exists()
    rows = _read_sheet(fp, "Data")
    assert rows[0] == ["ts", "v"]


def test_excel_make_naive_tz_aware_succeeds(tmp_path):
    # make_naive=True must strip tz from tz-aware columns so openpyxl can write
    # them (this is the whole point of the flag).
    aware = pd.to_datetime(["2020-01-01", "2020-01-02"]).tz_localize("UTC")
    df = pd.DataFrame({"ts": aware, "v": [1, 2]})
    fp = tmp_path / "out.xlsx"
    dfu.df_to_excel_worksheet(df, "Data", fp, make_naive=True)
    assert fp.exists()
    rows = _read_sheet(fp, "Data")
    assert rows[0] == ["ts", "v"]


def test_excel_chunking(tmp_path):
    df = pd.DataFrame({"a": list(range(10))})
    fp = tmp_path / "out.xlsx"
    dfu.df_to_excel_worksheet(df, "Data", fp, chunk_size=3)
    rows = _read_sheet(fp, "Data")
    # header + 10 data rows
    assert len(rows) == 11
    assert rows[0] == ["a"]
    assert rows[-1] == [9]


def test_make_df_naive_helper_direct_naive_passthrough():
    # A naive datetime column passes through unchanged (stays tz-naive, not
    # mangled).
    naive = pd.to_datetime(["2020-01-01", "2020-01-02"])
    df = pd.DataFrame({"ts": naive})
    out = dfu._make_df_naive(df)
    assert out["ts"].dt.tz is None


def test_make_df_naive_helper_strips_tz_aware():
    # _make_df_naive must drop the timezone from tz-aware columns.
    aware = pd.to_datetime(["2020-01-01"]).tz_localize("UTC")
    df = pd.DataFrame({"ts": aware})
    out = dfu._make_df_naive(df)
    assert out["ts"].dt.tz is None


# ---------------------------------------------------------------------------
# move_columns_far_left / move_columns_far_right
# ---------------------------------------------------------------------------

def test_move_columns_far_left():
    df = pd.DataFrame({"a": [1], "b": [2], "c": [3]})
    out = dfu.move_columns_far_left(df, ["c", "b"])
    assert list(out.columns) == ["c", "b", "a"]


def test_move_columns_far_left_empty():
    df = pd.DataFrame()
    out = dfu.move_columns_far_left(df, ["x"])
    assert out.empty


def test_move_columns_far_right():
    df = pd.DataFrame({"a": [1], "b": [2], "c": [3]})
    out = dfu.move_columns_far_right(df, ["a"])
    assert list(out.columns) == ["b", "c", "a"]


def test_move_columns_far_right_empty():
    df = pd.DataFrame()
    out = dfu.move_columns_far_right(df, ["x"])
    assert out.empty
